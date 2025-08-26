from typing import Optional, List, Dict, Any, cast
import os
import datetime
import argparse

try:
    import yaml  # type: ignore
except Exception:
    yaml = None

from utils import to_float, to_int
from income import compute_income_from_config
from expenses import compute_expenses_from_config
from depreciation import compute_depreciation_from_config
from cashflow import compute_cashflow_from_config
from loan import loan_schedule_annually

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
    from openpyxl.worksheet.worksheet import Worksheet
except Exception as e:
    raise RuntimeError("openpyxl is required. Install with 'pip install openpyxl'")


def _read_yaml(yaml_path: Optional[str]) -> str:
    if yaml_path and os.path.exists(yaml_path):
        return yaml_path
    # fallback to root project_config.yml
    from income import _find_project_config
    path = _find_project_config()
    if not path:
        raise FileNotFoundError("project_config.yml not found")
    return path


def _yen(n: float | int) -> int:
    try:
        return int(round(float(n)))
    except Exception:
        return 0


def export_excel(yaml_path: Optional[str] = None, out_path: Optional[str] = None, years: int = 40) -> str:
    """
    YAMLを読み込み、収入・支出・減価償却・キャッシュフロー・ローン残高をまとめたExcelを出力する。
    返り値: 出力したファイルパス
    """
    yml = _read_yaml(yaml_path)

    income_rows = compute_income_from_config(yaml_path=yml, years=years, round_to_yen=True)
    expense_rows = compute_expenses_from_config(yaml_path=yml, years=years, round_to_yen=True)
    depr = compute_depreciation_from_config(yaml_path=yml)
    cash = compute_cashflow_from_config(yaml_path=yml, years=years)

    # Prepare workbook
    wb = Workbook()
    ws: Worksheet = cast(Worksheet, wb.active)
    assert ws is not None
    ws.title = "Summary"

    # Styles
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.alignment = Alignment(horizontal="center")
    header.fill = PatternFill("solid", fgColor="F2F2F2")
    header.border = border

    yen = NamedStyle(name="yen")
    yen.number_format = "#,##0"
    yen.border = border

    pct = NamedStyle(name="pct")
    pct.number_format = "0.00%"
    pct.border = border

    # avoid duplicate named styles
    try:
        existing_names = {getattr(s, "name", str(s)) for s in wb.named_styles}
    except Exception:
        existing_names = set()
    for st in (header, yen, pct):
        if st.name not in existing_names:
            try:
                wb.add_named_style(st)
            except Exception:
                pass

    # Header row
    cols = [
        "年",
        "収入合計(円)",
        "支出合計(円)",
        "減価償却(円)",
        "キャッシュフロー(円)",
        "通算CF(円)",
        "ローン残高(円)",
        "売却価額(円)",
        "売却税(円)",
        "通算損益(円)",
        "APR(累積)",
    ]
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        ws.cell(row=1, column=c).style = header

    # Build depreciation per year (sum building+equipment, pad to years)
    b = depr.get("building", {}).get("annual_depreciation", []) or []
    e = depr.get("equipment", {}).get("annual_depreciation", []) or []
    dep_by_year: List[int] = []
    for i in range(years):
        bi = b[i] if i < len(b) else 0
        ei = e[i] if i < len(e) else 0
        dep_by_year.append(_yen(bi + ei))

    # Rows
    for i in range(years):
        inc = _yen(income_rows[i].get("annual_income", 0)) if i < len(income_rows) else 0
        exp = _yen(expense_rows[i].get("total_expenses", 0)) if i < len(expense_rows) else 0
        dep = dep_by_year[i] if i < len(dep_by_year) else 0
        cf = _yen(cash[i].get("cashflow", 0)) if i < len(cash) else (inc - exp)
        cum = _yen(cash[i].get("cashflow_cum", 0)) if i < len(cash) else 0
        bal = _yen(cash[i].get("loan_balance", cash[i].get("loan_balance_end", 0))) if i < len(cash) else 0
        sale = _yen(cash[i].get("sale_price", 0)) if i < len(cash) else 0
        tax_sale = _yen(cash[i].get("tax_on_sale", 0)) if i < len(cash) else 0
        net_profit = _yen(cash[i].get("net_profit", 0)) if i < len(cash) else 0
        apr = float(cash[i].get("apr", 0)) if i < len(cash) else 0.0
        ws.append([i + 1, inc, exp, dep, cf, cum, bal, sale, tax_sale, net_profit, apr])

    # Column widths & number formats
    widths = [6, 15, 15, 15, 16, 16, 16, 16, 14, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for r in range(2, 2 + years):
        for c in range(2, 11):  # monetary columns
            ws.cell(row=r, column=c).style = "yen"
        ws.cell(row=r, column=11).style = "pct"

    # Freeze header
    ws.freeze_panes = "A2"

    # Income sheet
    ws_inc: Worksheet = cast(Worksheet, wb.create_sheet("Income"))
    assert ws_inc is not None
    inc_cols = ["年", "月額賃料", "賃料変動率", "空室率", "年間総賃料", "年間収入"]
    ws_inc.append(inc_cols)
    for c in range(1, len(inc_cols) + 1):
        ws_inc.cell(row=1, column=c).style = header
    for i in range(min(years, len(income_rows))):
        r = income_rows[i]
        ws_inc.append([
            r.get("year", i + 1),
            _yen(r.get("monthly_rent", 0)),
            r.get("rent_change_rate", 0.0),
            r.get("vacancy_rate", 0.0),
            _yen(r.get("annual_gross", 0)),
            _yen(r.get("annual_income", 0)),
        ])
    for r in range(2, 2 + min(years, len(income_rows))):
        ws_inc.cell(row=r, column=2).style = "yen"
        ws_inc.cell(row=r, column=5).style = "yen"
        ws_inc.cell(row=r, column=6).style = "yen"
        ws_inc.cell(row=r, column=3).style = "pct"
        ws_inc.cell(row=r, column=4).style = "pct"

    # Expenses sheet
    ws_exp: Worksheet = cast(Worksheet, wb.create_sheet("Expenses"))
    assert ws_exp is not None
    exp_cols = [
        "年","固資(土)","都計(土)","固資(家)","都計(家)","税(合計)",
        "元金","利息","返済(合計)",
        "管理費","修繕費","保険料","光熱費","大規模修繕","設備修繕","運営(合計)",
        "支出合計"
    ]
    ws_exp.append(exp_cols)
    for c in range(1, len(exp_cols) + 1):
        ws_exp.cell(row=1, column=c).style = header
    for i in range(min(years, len(expense_rows))):
        t = expense_rows[i]
        ws_exp.append([
            t.get("year", i + 1),
            _yen(t.get("fixed_tax_land", 0)), _yen(t.get("city_tax_land", 0)),
            _yen(t.get("fixed_tax_building", 0)), _yen(t.get("city_tax_building", 0)), _yen(t.get("taxes_total", 0)),
            _yen(t.get("loan_principal", 0)), _yen(t.get("loan_interest", 0)), _yen(t.get("loan_total", 0)),
            _yen(t.get("management_fee", 0)), _yen(t.get("repairs", 0)), _yen(t.get("insurance", 0)), _yen(t.get("utilities", 0)),
            _yen(t.get("capex_large", 0)), _yen(t.get("equipment_repairs", 0)), _yen(t.get("operations_total", 0)),
            _yen(t.get("total_expenses", 0)),
        ])
    for r in range(2, 2 + min(years, len(expense_rows))):
        for c in range(2, len(exp_cols) + 1):
            ws_exp.cell(row=r, column=c).style = "yen"

    # Depreciation sheet
    ws_dep: Worksheet = cast(Worksheet, wb.create_sheet("Depreciation"))
    assert ws_dep is not None
    dep_cols = [
        "年",
        "減価償却(建物)",
        "減価償却(設備)",
        "減価償却(合計)",
        "減価償却累計",
        "簿価(建物)",
        "簿価(設備)",
        "簿価(合計)",
    ]
    ws_dep.append(dep_cols)
    for c in range(1, len(dep_cols) + 1):
        ws_dep.cell(row=1, column=c).style = header

    # Build per-year series and cumulative
    b_list = depr.get("building", {}).get("annual_depreciation", []) or []
    e_list = depr.get("equipment", {}).get("annual_depreciation", []) or []
    # Derive original costs assuming 1-yen residual policy: cost = sum(schedule)+1
    orig_b = _yen(sum(b_list) + 1)
    orig_e = _yen(sum(e_list) + 1)
    cum = 0
    cum_b = 0
    cum_e = 0
    for i in range(years):
        b_y = _yen(b_list[i] if i < len(b_list) else 0)
        e_y = _yen(e_list[i] if i < len(e_list) else 0)
        t_y = _yen(b_y + e_y)
        cum = _yen(cum + t_y)
        cum_b = _yen(cum_b + b_y)
        cum_e = _yen(cum_e + e_y)
        book_b = _yen(orig_b - cum_b)
        book_e = _yen(orig_e - cum_e)
        book_t = _yen(book_b + book_e)
        ws_dep.append([i + 1, b_y, e_y, t_y, cum, book_b, book_e, book_t])

    # Column widths & styles
    dep_widths = [6, 18, 18, 18, 18, 14, 14, 14]
    for i, w in enumerate(dep_widths, start=1):
        ws_dep.column_dimensions[chr(64 + i)].width = w
    for r in range(2, 2 + years):
        for c in range(2, len(dep_cols) + 1):
            ws_dep.cell(row=r, column=c).style = "yen"
    ws_dep.freeze_panes = "A2"

    # Loan sheet (annual)
    # Derive loan params from YAML
    with open(yml, "r", encoding="utf-8") as fh:
        if yaml is None:
            raise RuntimeError("PyYAML is required to load YAML")
        yaml_mod = cast(Any, yaml)
        cfg = yaml_mod.safe_load(fh) or {}
    purchase_price = to_float(cfg.get("purchase_price", 0.0), 0.0)
    initial_capital_ratio = to_float(cfg.get("initial_capital_ratio", 0.2), 0.2)
    principal = purchase_price * (1.0 - initial_capital_ratio)
    annual_rate = to_float(cfg.get("annual_rate", cfg.get("loan", {}).get("annual_rate", 0.0)), 0.0)
    loan_years = to_int(cfg.get("years", cfg.get("loan", {}).get("years", years)), years)
    start_month = to_int(cfg.get("start_month", cfg.get("loan", {}).get("start_month", 1)), 1)
    method = str(cfg.get("method", cfg.get("loan", {}).get("method", "equal_total")))
    group_by = str(cfg.get("group_by", cfg.get("loan", {}).get("group_by", "calendar")))
    sched = loan_schedule_annually(principal, annual_rate, loan_years, start_month, method, group_by)
    annual = sched.get("annual", [])

    ws_loan: Worksheet = cast(Worksheet, wb.create_sheet("Loan"))
    assert ws_loan is not None
    loan_cols = ["年", "月数", "元金", "利息", "返済(合計)", "累計返済", "期末残高"]
    ws_loan.append(loan_cols)
    for c in range(1, len(loan_cols) + 1):
        ws_loan.cell(row=1, column=c).style = header
    for i in range(min(years, len(annual))):
        a = annual[i]
        ws_loan.append([
            a.get("year_index", i + 1), a.get("months", 12),
            _yen(a.get("principal_paid", 0)), _yen(a.get("interest_paid", 0)), _yen(a.get("total_paid", 0)),
            _yen(a.get("cumulative_paid", 0)), _yen(a.get("balance_end", 0)),
        ])
    for r in range(2, 2 + min(years, len(annual))):
        for c in range(3, 8):
            ws_loan.cell(row=r, column=c).style = "yen"

    # Config sheet (raw YAML values)
    ws_cfg: Worksheet = cast(Worksheet, wb.create_sheet("Config"))
    assert ws_cfg is not None
    ws_cfg.append(["Key", "Value"])
    ws_cfg.cell(row=1, column=1).style = header
    ws_cfg.cell(row=1, column=2).style = header
    # dump key-values (shallow)
    for k, v in (cfg or {}).items():
        ws_cfg.append([str(k), str(v)])

    # Save
    if out_path is None:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(os.path.dirname(yml), f"plan_{ts}.xlsx")
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export real estate plan to Excel")
    parser.add_argument("--config", "-c", dest="config", default=None, help="Path to project_config.yml")
    parser.add_argument("--out", "-o", dest="out", default=None, help="Output .xlsx path")
    parser.add_argument("--years", "-y", dest="years", type=int, default=40, help="Number of years")
    args = parser.parse_args()
    path = export_excel(args.config, args.out, args.years)
    print(f"Excel written: {path}")
